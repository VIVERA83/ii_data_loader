from io import BytesIO

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from store.excel.exeptions import ExceptionExcelFormat


async def excel_file_handler(file: BytesIO) -> list[tuple[str, str, str]]:
    """Returns list of tuple strings [username, login, password]."""
    wb: Workbook = load_workbook(file)
    try:
        return [
            (user.value, login.value, password.value)
            for user, login, password in wb[wb.sheetnames[0]]
        ]
    except ValueError:
        raise ExceptionExcelFormat()
